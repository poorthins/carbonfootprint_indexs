import openpyxl
import pandas as pd

project_file_path = "C:\\Users\\Tim\\Desktop\\python\\論文\\工程計算\\論文工程資料\\碳足跡\\數據品質\\工程數據品質_v2\\20.xlsx"
quality_file_path = "碳足跡\\數據品質\\數據品質係數_v1\\品質_20.xlsx"

wb = openpyxl.load_workbook(project_file_path)

#在標單詳細表中計算工程總碳排
dic_sheet = wb['標單詳細表']
dic_sheet['I1'] = '工程總碳排'
total_carbon_emission = sum(cell.value for cell in dic_sheet['H'] if isinstance(cell.value, (int, float)))
dic_sheet['I2'] = total_carbon_emission

#刪除上次計算結果
second_sheet = wb['單價分析表']
second_sheet.delete_cols(9) #刪除第I column
second_sheet.delete_cols(10) #刪除第J column
second_sheet.delete_cols(11) #刪除第K column
second_sheet.delete_cols(12) #刪除第L column
second_sheet.delete_cols(13) #刪除第M column

second_sheet['I1'] = '材料總碳排'
for row in range(2, second_sheet.max_row + 1):
    if second_sheet[f'E{row}'].value is not None and second_sheet[f'H{row}'].value is not None:
        second_sheet[f'I{row}'] = second_sheet[f'E{row}'].value * second_sheet[f'H{row}'].value

#計算Fi值
second_sheet['J1'] = 'Fi'
for row in range(2, second_sheet.max_row + 1):
    if second_sheet[f'I{row}'].value is not None and second_sheet[f'A{row}'].value is None:
        second_sheet[f'J{row}'] = (second_sheet[f'I{row}'].value / total_carbon_emission) * 100

wb.save(project_file_path) 

#將DQR分數放上Excel sheet
quality_df = pd.read_excel(quality_file_path, usecols=['編碼', 'DQR'])
project_df = pd.read_excel(project_file_path, sheet_name='單價分析表')

# 進行excel Vlookup的操作
for index, row in project_df.iterrows():
    if row['編碼'] in quality_df['編碼'].values:
        dqr_value = quality_df.loc[quality_df['編碼'] == row['編碼'], 'DQR'].values
        if dqr_value.size > 0:
            project_df.at[index, 'DQR'] = dqr_value[0]

#計算DQRw
project_df['DQRw'] = pd.to_numeric(project_df['DQR'], errors='coerce') * pd.to_numeric(project_df['Fi'], errors='coerce')      

with pd.ExcelWriter(project_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    project_df.to_excel(writer, sheet_name='單價分析表', index=False)
 
dqrw_accumulated = 0
fi_accumulated = 0
last_non_zero_item_index = None
appendix_data = [] #補appendix缺失的bug

# 第一回計算工程 DQRw 和 Fi
for index, row in project_df.iterrows():
    
    if pd.notna(row['項次']) :
        last_non_zero_item_index = index

    if row['名稱'] == '合計' and last_non_zero_item_index is not None:
        if fi_accumulated != 0 :
            ratio = (dqrw_accumulated / fi_accumulated)
            #project_df.at[last_non_zero_item_index, 'DQRtotal'] = ratio
            appendix_data.append((project_df.at[last_non_zero_item_index, '編碼'], fi_accumulated, dqrw_accumulated))
        dqrw_accumulated = 0
        fi_accumulated = 0
        last_non_zero_item_index = None
    dqrw_accumulated += row.get('DQRw', 0) if pd.notna(row.get('DQRw', 0)) else 0
    fi_accumulated += row.get('Fi', 0) if pd.notna(row.get('Fi', 0)) else 0

with pd.ExcelWriter(project_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    project_df.to_excel(writer, sheet_name='單價分析表', index=False)
# 將appendix的數據品質補充回去excel中
for index, row in project_df.iterrows():
    # 檢查條件：項次為NaN，Fi有數字但DQRw為NaN
    if pd.isna(row['項次']) and pd.notna(row['Fi']) and pd.isna(row['DQRw']):
        # 在 appendix_data 中尋找相同編碼的條目
        for entry in appendix_data:
            code, fi_value, dqrw_value = entry
            if code == row['編碼']:
                project_df.at[index, 'Fi'] = fi_value
                project_df.at[index, 'DQRw'] = dqrw_value
                break 

# 最後再算一次數據品質
m=[]
for index, row in project_df.iterrows():
    if pd.notna(row['項次']) and row['項次'] != 0:
        last_non_zero_item_index = index

    if pd.notna(row['名稱']) and row['名稱'] == '合計' and last_non_zero_item_index is not None:
            # 計算比例並記錄結果
        if fi_accumulated != 0:
            ratio = (dqrw_accumulated / fi_accumulated)
            project_df.at[last_non_zero_item_index, 'Fi'] = fi_accumulated
            project_df.at[last_non_zero_item_index, 'DQRw'] = dqrw_accumulated
            project_df.at[last_non_zero_item_index, 'DQRtotal'] = ratio
            m.append(ratio)
            # 重置累計值
        dqrw_accumulated = 0
        fi_accumulated = 0
        last_non_zero_item_index = None
        # 累計 DQRw 和 Fi
    dqrw_accumulated += row.get('DQRw', 0) if pd.notna(row.get('DQRw', 0)) else 0
    fi_accumulated += row.get('Fi', 0) if pd.notna(row.get('Fi', 0)) else 0
mean = round(sum(m)/len(m),2)
# 計算數據品質結果
print(mean)
with pd.ExcelWriter(project_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    project_df.to_excel(writer, sheet_name='單價分析表', index=False)

wb = openpyxl.load_workbook(project_file_path)
dic_sheet = wb['標單詳細表']
dic_sheet['J1'] = "數據品質總平均"
dic_sheet['K1'] = "整體數據品質水平"
dic_sheet['J2'] = mean
if mean <= 1.7:
    dic_sheet['K2'] = "高品質"
elif 1.7< mean <= 3:
    dic_sheet['K2'] = "基本品質"
else:
    dic_sheet['K2'] = "初估品質"
print(dic_sheet['K2'].value)
wb.save(project_file_path)
