import openpyxl
import numpy as np

file_path = 'lee_sheet.xlsx'

#step1 初步計算部門百分比
def calculate_percentages(sheet, sheet_487, department_row, item_row, value_row, percentage_row, start_col):
    
    # 判斷部門數量
    end_col = start_col
    while sheet.cell(row=department_row, column=end_col).value is not None:
        end_col += 1
    num_items = end_col - start_col #部門數量

    # 紀錄部門相關數據
    departments = [sheet.cell(row=department_row, column=col).value for col in range(start_col, end_col)]
    items = [sheet.cell(row=item_row, column=col).value for col in range(start_col, end_col)]
    values = [sheet.cell(row=value_row, column=col).value if sheet.cell(row=value_row, column=col).value is not None else 0 for col in range(start_col, end_col)]
    total_value_sum = sum(values)
    data = {}
    percentage = {}
    
    #計算部門銷售值數據
    if department_row == 9:
        
        #計算各部門金額
        department_sums = {}
        for department, value in zip(departments, values):
            if department not in department_sums:
                department_sums[department] = 0
            department_sums[department] += value

        # 計算各部門金額占比
        for department, item, value in zip(departments, items, values):
            if department not in data:
                data[department] = {
                    'items': [],
                    'values': [],
                    'percentages': []
                }
            if department_sums[department] == value:  # 僅一個部門狀況
                percentage = 100
            else:
                percentage = (value / department_sums[department]) * 100
            data[department]['items'].append(item)
            data[department]['values'].append(value)
            data[department]['percentages'].append(percentage)

        # 將數據寫入excel
        for col, department in enumerate(departments, start=start_col):
            percentage = data[department]['percentages'][data[department]['items'].index(items[col - start_col])]
            sheet.cell(row=percentage_row, column=col, value=percentage)
    
    # 計算部門生產值數據
    if department_row == 4:

        #設定初始參數
        check_items = False
        name1 = sheet['B1'].value
        denominator_value = 0
        numerator = 0

        #設定對照字典
        dic = {}
        for department, item, value in zip(departments, items, values):
            if department not in dic:
                dic[department] = {
                    'items': [],
                    'value': [],
                }
            dic[department]['items'].append(item)
            dic[department]['value'].append(value)
        #讀取487部門頁面並處理要拆分的部門
        for row_index, row in enumerate(sheet_487.iter_rows(min_row=1, max_row=sheet_487.max_row), start=1):
            if row[0].value == name1:
                department = row[1].value
                if department in dic:
                    if department not in data:
                        data[department] = {
                            'items': [],
                            'values': [],
                            'percentages': []
                        }
                    for item, value in zip(dic[department]['items'], dic[department]['value']):
                        data[department]['items'].append(item)
                        percent = (value / total_value_sum) * 100
                        data[department]['percentages'].append(percent)
                        data[department]['values'].append(sheet_487.cell(row=row_index, column=176).value)
                    denominator_value += sheet_487.cell(row=row_index, column=176).value
                    numerator += sheet_487.cell(row=row_index, column=176).value
                else: 
                    department = row[1].value
                    if department not in data:
                        data[department] = {
                            'items': [],
                            'values': [],
                            'percentages': []
                        }
                    data[department]['items'].append(row[2].value)
                    data[department]['values'].append(sheet_487.cell(row=row_index, column=176).value)
                    data[department]['percentages'].append('prep')
                    denominator_value += sheet_487.cell(row=row_index, column=176).value
                    check_items = True
        fraction_num =  (numerator/denominator_value)*100

        #工業產銷存的部門與487部門的關係
        if not check_items:
            for col, department in enumerate(departments, start=start_col):
                percentage = data[department]['percentages'][data[department]['items'].index(items[col - start_col])]
                sheet.cell(row=percentage_row, column=col, value=percentage)
        else:
            for department in data:
                for i, item in enumerate(data[department]['items']):
                    if isinstance(data[department]['percentages'][i], (int, float)):
                        data[department]['percentages'][i] = data[department]['percentages'][i] * fraction_num / 100
                    elif data[department]['percentages'][i] == 'prep':
                        data[department]['percentages'][i] = (data[department]['values'][i] / denominator_value) * 100
                        
    return num_items, data

#執行step1
def process_file(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['控制台']
    sheet_487 = wb['487']

    # 計算銷售值
    num2, data2 = calculate_percentages(sheet, sheet_487, department_row=9, item_row=10, value_row=11, percentage_row=12, start_col=2)

    # 計算生產值
    num1, data1 = calculate_percentages(sheet, sheet_487, department_row=4, item_row=5, value_row=6, percentage_row=7, start_col=2)
    wb.save(file_path)
    return num1, data1, num2, data2

#step2 執行拆分部門
def distribution_part(file_path, num1, data1, data2):
    #準備作業
    wb = openpyxl.load_workbook(file_path)
    control_sheet = wb['控制台']
    original_sheet = wb['164']
    copied_sheet = wb.copy_worksheet(original_sheet)

    if 'calculate_sheet' in wb.sheetnames: #每次計算都重新開始新的頁面
        del wb['calculate_sheet']    
    copied_sheet.title = 'calculate_sheet'
    
    #尋找487分頁中要拆分的對象後存取資料
    name1 = control_sheet['B1'].value 
    sheet_487 = wb['487']
    rows_to_insert = [] #存放在487分頁中找到的部門數據
    for row in sheet_487.iter_rows(min_row=1, max_row=sheet_487.max_row, min_col=1, max_col=sheet_487.max_column):
        if row[0].value == name1:
            rows_to_insert.append([cell.value for cell in row])
    records_bc = []
    count_department_num =0

    #Step 1:處理供給部門數據 
    for row in range(1, copied_sheet.max_row + 1):
        if copied_sheet.cell(row=row, column=2).value == name1:
            copied_sheet.delete_rows(row)
            #插入部門數量的row
            copied_sheet.insert_rows(row, num1)
            current_row = row
            
            for i, inserted_row in enumerate(rows_to_insert):
                inserted_department = inserted_row[1]
                #如果是有在工業產銷存部門則去rows_to_insert抓資料
                if inserted_department in data2:
                    num_items = len(data2[inserted_department]['items'])
                    for k in range(num_items):
                        count_department_num += 1
                        item = data2[inserted_department]['items'][k]
                        percentage = data2[inserted_department]['percentages'][k]
                        copied_sheet.cell(row=current_row + k, column=2, value=inserted_department)
                        copied_sheet.cell(row=current_row + k, column=1, value=percentage)
                        copied_sheet.cell(row=current_row + k, column=3, value=item)
                        inserted_value_c = copied_sheet.cell(row=current_row + k, column=3).value
                        records_bc.append(inserted_value_c)
                        for j, value in enumerate(inserted_row[3:], start=4):
                            copied_sheet.cell(row=current_row + k, column=j, value=value * percentage / 100)
                    current_row += num_items
                
                #沒有的話則建立新資料
                else:
                    copied_sheet.insert_rows(current_row, 1)
                    count_department_num += 1
                    for j, value in enumerate(inserted_row, start=1):
                        copied_sheet.cell(row=current_row, column=j, value=value)
                    current_row += 1
                    inserted_value_c = copied_sheet.cell(row=current_row, column=3).value
                    records_bc.append(inserted_value_c)
            break

    #Step 2:處理需求部門
    rows_to_insert1 = []
    for col in copied_sheet.iter_cols(min_row=1, max_row=copied_sheet.max_row, min_col=1, max_col=copied_sheet.max_column):
        if col[2].value == name1:
            rows_to_insert1.append([cell.value for cell in col])
    for col in range(1, copied_sheet.max_column + 1):
        if copied_sheet.cell(row=3, column=col).value == name1:
            copied_sheet.delete_cols(col)
            copied_sheet.insert_cols(col, count_department_num)
            current_col = col  # 紀錄位置
            for department in data1:
                for i, item in enumerate(data1[department]['items']):
                    copied_sheet.cell(row=4, column=current_col + i, value=item)
                for i, percentage in enumerate(data1[department]['percentages']):
                    copied_sheet.cell(row=3, column=current_col + i, value=percentage)
                    for j, inserted_row in enumerate(rows_to_insert1):
                        for k, value in enumerate(inserted_row[4:], start=5):
                            copied_sheet.cell(row=k, column=current_col + i, value=value * percentage / 100)
                current_col += len(data1[department]['items'])  # 更新 current_col
            break
                            
    wb.save(file_path)
    return count_department_num #紀錄總共有幾個部門

#Step 3:計算能源比例
def distribution_energy(file_path, num1, data1):
    #初始化資料
    wb = openpyxl.load_workbook(file_path)
    control_sheet = wb['控制台']
    b1_value = control_sheet['B1'].value
    if 'energy_distribution' in wb.sheetnames:
        del wb['energy_distribution']
    energy_sheet = wb['energy']
    copied_sheet = wb.copy_worksheet(energy_sheet)
    copied_sheet.title = 'energy_distribution'
    target_columns = []

    # 搜尋第三列，如果有一樣的值則將整行儲存做為目標行後刪除此行，並插入要拆分的行數
    for col in range(1, copied_sheet.max_column + 1):
        if copied_sheet.cell(row=3, column=col).value == b1_value:
            column_data = [copied_sheet.cell(row=row, column=col).value for row in range(1, copied_sheet.max_row + 1)]
            target_columns.append(column_data)
            copied_sheet.delete_cols(col)
            copied_sheet.insert_cols(col, num1) #插入拆分比例
            break  # 只會有一列匹配

    #開始計算能源比例
    current_col = col
    for department in data1:
        for i, item in enumerate(data1[department]['items']):
            copied_sheet.cell(row=4, column=current_col + i, value=item)
        
        for i, percentage in enumerate(data1[department]['percentages']):
            copied_sheet.cell(row=3, column=current_col + i, value=percentage)
            copied_sheet.cell(row=2, column=current_col + i, value=b1_value)

            for row in range(5, copied_sheet.max_row + 1):  # 將percentages乘以剛剛刪除的目標行   
                if target_columns[0][row - 1] is not None:
                    target_value = target_columns[0][row - 1]
                    copied_sheet.cell(row=row, column=current_col + i, value=(target_value * percentage / 100 if target_value is not None else 0))
        current_col += len(data1[department]['items'])

    wb.save(file_path)            
    energy_percentage = [copied_sheet.cell(row=35, column=c).value for c in range(1, copied_sheet.max_column + 1)]

    return energy_percentage

#step 4:計算新能源平衡表
def new_energy_balance(file_path, num1, data1,energy_percentage):
    wb = openpyxl.load_workbook(file_path)
    control_sheet = wb['控制台']
    b1_value = control_sheet['B1'].value

    #複製new_energy_balance這個分頁，並取名calculate_new_energy
    if 'calculate_new_energy' in wb.sheetnames:
        del wb['calculate_new_energy']
    original_sheet = wb['new_energy_balance']
    copied_sheet = wb.copy_worksheet(original_sheet)
    copied_sheet.title = 'calculate_new_energy'

    #搜尋D row然後如果有遇到一樣的值則將整個行向下複製num1數量的行數
    for row in range(1, copied_sheet.max_row + 1):
        if copied_sheet.cell(row=row, column=4).value == b1_value:
            #將data1中的item與percentage放回去向下複製的行數
            copied_sheet.insert_rows(row + 1, num1)
            for i in range(num1):
                for col in range(1, copied_sheet.max_column + 1):
                    copied_sheet.cell(row=row + 1 + i, column=col, value=copied_sheet.cell(row=row, column=col).value)
            
            current_row = row + 1
            for department in data1:
                for i, item in enumerate(data1[department]['items']):
                    copied_sheet.cell(row=current_row + i, column=5, value=item)
                    copied_sheet.cell(row=current_row + i, column=3, value=copied_sheet.cell(row=row, column=3).value)
                    copied_sheet.cell(row=current_row + i, column=4, value=copied_sheet.cell(row=row, column=4).value)
                for i, percentage in enumerate(data1[department]['percentages']):
                    copied_sheet.cell(row=current_row + i, column=6, value=percentage)  # 假设百分比放在第6列
                current_row += len(data1[department]['items'])
            
            copied_sheet.delete_rows(row)
            break          
    
    # 將energy_percentage填到這個頁面的第AR列，從第4行開始填
    copied_sheet.delete_cols(44)
    copied_sheet.insert_cols(44, 1)
    for i, percentage in enumerate(energy_percentage, start=3):
        copied_sheet.cell(row=i, column=44, value=percentage)  # AR列是第44列

    # 讓同行的AT=AR*AQ
    copied_sheet.delete_cols(47)
    copied_sheet.insert_cols(47, 1)
    for i in range(6, copied_sheet.max_row + 1):
        ar_value = copied_sheet.cell(row=i, column=44).value
        aq_value = copied_sheet.cell(row=i, column=43).value  # AQ列是第43列
        as_value = copied_sheet.cell(row=i, column=45).value
        if as_value is None:
            copied_sheet.cell(row=i, column=46, value=ar_value * aq_value)  # AT列是第46列
    
    # 搜尋並紀錄國內生產總額數據
    calculate_sheet = wb['calculate_sheet']
    col_index = None
    for col in range(1, calculate_sheet.max_column + 1):
        if calculate_sheet.cell(row=4, column=col).value == '國內生產總額':
            col_index = col
            break

    if col_index is not None:
        for row in range(4, calculate_sheet.max_row ):
            value = calculate_sheet.cell(row=row, column=col_index).value
            copied_sheet.cell(row=row+1, column=47, value=value)

    # 計算能源係數
    energy_index=[]
    for row in range(6, copied_sheet.max_row + 1):
        at_value = copied_sheet.cell(row=row, column=46).value  
        au_value = copied_sheet.cell(row=row, column=47).value     
        if  au_value == 0 or at_value == 0:
            result =0
            copied_sheet.cell(row=row, column=48, value=result)
            energy_index.append(result)
        elif isinstance(at_value, (int, float)) and isinstance(au_value, (int, float)):
            result = at_value / au_value / 1000000
            copied_sheet.cell(row=row, column=48, value=result)
            energy_index.append(result)
    wb.save(file_path)
    return energy_index

#step5 計算矩陣
def extract_matrix(file_path,energy_index):
    wb = openpyxl.load_workbook(file_path)
    calculate_sheet = wb['calculate_sheet']

    # 讀取C列與D列，找到開始和結束的位置
    start_row = None
    end_row = None
    for row in range(1, calculate_sheet.max_row + 1):
        c_value = calculate_sheet.cell(row=row, column=3).value
        d_value = calculate_sheet.cell(row=row, column=4).value
        if isinstance(c_value, str) and isinstance(d_value, (int, float)):
            if start_row is None:
                start_row = row
        if c_value == '中間投入合計':
            end_row = row - 1
            break

    # 讀取第四行與第五行，找到開始和結束的位置
    start_col = None
    end_col = None
    gdp_col = None
    gdp_data = []

    for col in range(1, calculate_sheet.max_column + 1):
        row4_value = calculate_sheet.cell(row=4, column=col).value
        row5_value = calculate_sheet.cell(row=5, column=col).value
        if isinstance(row4_value, str) and isinstance(row5_value, (int, float)):
            if start_col is None:
                start_col = col
        if row4_value == '中間需要合計':
            end_col = col - 1

        if row4_value == '國內生產總額':
            gdp_col = col
            gdp_data = [calculate_sheet.cell(row=row, column=gdp_col).value for row in range(start_row, end_row + 1)]

    if '1-A' in wb.sheetnames:
        del wb['1-A']
    
    new_sheet = wb.create_sheet(title='1-A')
    
    # 初始化原始矩阵
    original_matrix = []

    # 計算A矩陣
    for row in range(start_row, end_row + 1):
        matrix_row = []
        for col in range(start_col, end_col + 1):
            value = calculate_sheet.cell(row=row, column=col).value
            gdp_value = gdp_data[col - start_col]
            if isinstance(value, (int, float)) and isinstance(gdp_value, (int, float)) and gdp_value != 0:
                value = value / gdp_value
            matrix_row.append(value)
            new_sheet.cell(row=row - start_row + 1, column=col - start_col + 1, value=value)
        original_matrix.append(matrix_row)

    original_matrix = np.array(original_matrix)

    # 計算inverse(1-A)
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    identity_matrix = np.eye(num_rows, num_cols)

    #1-A
    result_matrix = identity_matrix - original_matrix

    #inv(1-A)
    inverse_matrix = np.linalg.inv(result_matrix)

    # 將INV(1-A)寫入EXCEL中
    start_new_row = end_row + 2
    for i in range(num_rows):
        for j in range(num_cols):
            new_sheet.cell(row=start_new_row + i, column=1+j, value=inverse_matrix[i, j])
    
    for i, value in enumerate(energy_index, start=1):
        new_sheet.cell(row=start_new_row + num_rows + 5, column=i, value=value)
    
    mmult_result = np.dot(energy_index, inverse_matrix)

    index = []
    # 計算矩陣

    for i, value in enumerate(mmult_result, start=1):
        new_sheet.cell(row=start_new_row + num_rows + 7, column=i, value=value)
        index.append(float(value))
    
    wb.save(file_path)
    return index

#step6 產出結果
def process_supply_section(file_path, index, data1):
    wb = openpyxl.load_workbook(file_path)
    
    # 记录 "供给部门" 和 "中间需要合计" 之间的位置
    calculate_sheet = wb['calculate_sheet']

    start_col = None
    end_col = None

    for col in range(1, calculate_sheet.max_column + 1):
        cell_value = calculate_sheet.cell(row=4, column=col).value
        if cell_value == "供給部門":
            start_col = col + 1  
        elif cell_value == "中間需要合計" and start_col is not None:
            end_col = col - 1  
            break

    # 复制本次部门内容
    row_data = []
    for col in range(start_col, end_col + 1):
        row_data.append(calculate_sheet.cell(row=4, column=col).value)
    
    # 创建一个final_sheet并将结果放进去
    if 'final_sheet' in wb.sheetnames:
        del wb['final_sheet']

    final_sheet = wb.create_sheet(title='final_sheet')

    for col_index, value in enumerate(row_data, start=1):
        final_sheet.cell(row=col_index, column=2, value=value)

    for row_index, value in enumerate(index, start=1):
        final_sheet.cell(row=row_index, column=3, value=value)
        department = calculate_sheet.cell(row=4 + row_index, column=2).value
        final_sheet.cell(row=row_index, column=1, value=department)
        # 如果遇到 data1 中的 department，则覆盖 data1 中的 value
        if department in data1:
            for item_index, item in enumerate(data1[department]['items']):
                if item == final_sheet.cell(row=row_index, column=2).value:
                    data1[department]['values'][item_index] = value
    wb.save(file_path)

#step7 更新計算紀錄
def update_calculation_records(file_path, data1, count_department_num):
    wb = openpyxl.load_workbook(file_path)
    control_sheet = wb['控制台']
    b1_value = control_sheet['B1'].value
    record_sheet = wb['紀錄用']

    for row in range(1, record_sheet.max_row + 1):
        if record_sheet.cell(row=row, column=1).value == b1_value:
            record_sheet.delete_rows(row)
            record_sheet.insert_rows(row, count_department_num)
            # 插入新的记录
            current_row = 1
            for department, data in data1.items():
                for item, value in zip(data['items'], data['values']):
                    while record_sheet.cell(row=current_row, column=1).value is not None:
                        current_row += 1
                    record_sheet.cell(row=current_row, column=1, value=department)
                    record_sheet.cell(row=current_row, column=2, value=item)
                    record_sheet.cell(row=current_row, column=3, value=value)
                    current_row += 1
        else:
            for department, data in data1.items():
                for item, value in zip(data['items'], data['values']):
                    if record_sheet.cell(row=row, column=2).value == item:
                        record_sheet.cell(row=row, column=3, value=value)

    wb.save(file_path)

def batch_process(file_path, targets):
    
    
    for target_name in targets:
        wb = openpyxl.load_workbook(file_path)
        original_control_sheet = wb['控制台']
        new_control_sheet_name = original_control_sheet.cell(row=1, column=2).value
        if isinstance(new_control_sheet_name, int):
            new_control_sheet_name = str(new_control_sheet_name)
        
        # Save the current workbook to apply any pending changes
        original_control_sheet.title = new_control_sheet_name

        target_sheet = wb[target_name]
        target_sheet.title = '控制台'
        wb.save(file_path)
        
        # 執行所有步驟
        num1, data1, num2, data2 = process_file(file_path)
        count_department_num = distribution_part(file_path, num1, data1, data2)
        energy_percentage = distribution_energy(file_path, count_department_num, data1)
        energy_index = new_energy_balance(file_path, count_department_num, data1, energy_percentage)
        index = extract_matrix(file_path, energy_index)
        process_supply_section(file_path, index, data1)
        update_calculation_records(file_path, data1, count_department_num)  


targets = ['12','14','15','19']
batch_process(file_path, targets)
#all ['13','30','35','42','43','44','45','47','49','50','51','52','53','56','57','59','62','63','66','67','69','70','71','73','75','76','80','81','87','89','90','91','92','94','96','97']
